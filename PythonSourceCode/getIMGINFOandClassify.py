# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
import subprocess
from datetime import datetime
from typing import List, Dict, Any, Union 
import re 
from tqdm import tqdm 
# import shutil # 移除非必要的导入，遵循难度等级1
# TODO tag的相关性分析
# TODO tag的统计分析
# TODO 读写图片信息PNGINFO，PNG转格式JPG(jpeg)/WEBP，压缩以前的PNG图片
# TODO windows自带图片评分功能
# TODO windows自带图片备注功能
# TODO 图片去重功能，webp和png格式转换后去重
# TODO 图片按评分排序功能
# TODO 图片按备注排序功能
# TODO 图片按标签排序功能
# TODO 图片按创建时间排序功能
# TODO 图片按修改时间排序功能
# TODO 图片按访问时间排序功能
# TODO 图片按文件大小排序功能
# 不写代码
# 优先考虑只使用windows的情况下，不需要刻意做适配其他系统
# TODO 获取文件的创建时间和修改时间，进行对比时间，哪个更早，哪个就是真实的创建时间
# TODO 将真实的创建时间写入图片的文件名之中，文件名格式为create-YYYYMMDD-原文件名.ext
# TODO A. 利用 pywin32 库并调用 Windows Shell API 是在 Windows 环境下批量修改文件属性以实现日期信息冗余存储的最优方案。【选择方案】
# TODO B. 整合 EXIF 方案： 将此 Windows Shell API 方案与之前讨论的 EXIF 元数据写入方案整合，实现双重冗余保护，确保日期万无一失。【不选择】【不想因为这个操作导致有问题】
# TODO 以上功能，读取和写入的函数都要分开来写，解耦


# TODO 要有测试函数
# TODO 检索硬盘找到创建时间和修改时间不一样，可以作为试验素材的文件，并将其复制到试验文件夹
# TODO 测试函数，测试读取和写入创建时间和修改时间的函数
# TODO 进行多次测试，确保读取和写入的函数都没有问题



# [2025-10-28] 新增: 导入图片扫描和元数据提取模块
try:
    # 导入核心扫描函数和日志函数
    from image_scanner import get_image_info, log_error
except ImportError:
    print("严重警告: 找不到核心功能模块 'image_scanner.py'。程序无法运行。")
    # 定义占位函数，确保主程序可以安全退出
    def log_error(message: str):
        print(f"[CRITICAL ERROR] image_scanner.py 缺失。{message}")
    def get_image_info(folder_path: str):
        print("[CRITICAL ERROR] image_scanner.py 缺失，无法扫描图片。")
        return []

# [2025-10-27] 新增: 导入 TF-IDF 核心功能
# 假设 tfidf_processor.py 文件存在于同一目录下
try:
    # 新增: 导入 format_tfidf_tags_for_filename 函数，用于生成文件后缀
    from tfidf_processor import preprocess_tags, calculate_and_extract_tfidf, format_tfidf_tags_for_filename
    # 配置 TF-IDF (注意：未来应从 tfidf_processor 导入这些常量以实现统一)
    TFIDF_NEW_COLUMN_NAME = 'TF-IDF区分度关键词(Top 10)'
    TFIDF_TOP_N_FEATURES = 10 
    TFIDF_TARGET_COLUMN = '提取正向词的核心词'
    TFIDF_SUFFIX_COLUMN = 'TF-IDF文件名后缀' # 新增: 用于存储 TF-IDF 后缀的临时列名
except ImportError:
    print("警告: 找不到核心功能模块 'tfidf_processor.py'。TF-IDF 分析功能将被跳过。")
    # 定义占位函数和配置，避免程序崩溃
    TFIDF_NEW_COLUMN_NAME = 'TF-IDF区分度关键词(Top 10)'
    TFIDF_TOP_N_FEATURES = 10
    TFIDF_TARGET_COLUMN = '提取正向词的核心词'
    TFIDF_SUFFIX_COLUMN = 'TF-IDF文件名后缀'
    def preprocess_tags(tags_series): return [], tags_series 
    def calculate_and_extract_tfidf(df, corpus, cleaned_tags_series, top_n): return ["TF-IDF模块缺失"] * len(df), [[]] * len(df)
    def format_tfidf_tags_for_filename(tag_list, tag_delimiter="___"): return ""
# [2025-10-27] End TF-IDF 导入

# [2025-10-28] 新增: 导入文件名处理功能模块
try:
    # 导入文件名标记的核心函数和 TAGGING_KEYWORDS 常量
    from filename_tagger import tag_files_by_prompt, TAGGING_KEYWORDS 
except ImportError:
    print("警告: 找不到核心功能模块 'filename_tagger.py'。文件名标记功能将受限或失败。")
    # 定义占位函数和常量，避免程序崩溃
    def tag_files_by_prompt(image_data: List[Dict[str, Any]], keyword_list: List[str], tag_delimiter: str = "___", tfidf_suffix_col: str = None) -> List[Dict[str, Any]]:
        print("[ERROR] filename_tagger.py 缺失，跳过文件名标记。")
        return image_data
    TAGGING_KEYWORDS = [] # 占位空列表
# [2025-10-28] End filename_tagger 导入

# [2025-10-28] 新增: 导入文件分类和归档功能模块
try:
    # 导入文件分类和归档的核心函数
    from file_categorizer import categorize_images, archive_date_folders 
except ImportError:
    print("警告: 找不到核心功能模块 'file_categorizer.py'。文件分类和归档功能将被跳过。")
    # 定义占位函数，避免程序崩溃
    def categorize_images(image_data: List[Dict[str, Any]], keyword_list: List[str], root_dir_path: str):
        print("[ERROR] file_categorizer.py 缺失，跳过文件分类。")
    def archive_date_folders(base_source_dir: str, archive_target_dir: str):
        print("[ERROR] file_categorizer.py 缺失，跳过日期文件夹归档。")
# [2025-10-28] End file_categorizer 导入

# [2025-10-27] 新增: 导入 image_scorer_supervised 模块，用于添加评分列
try:
    from image_scorer_supervised import ScorerConfig, ImageScorer
    SCORER_MODULE_LOADED = True
except ImportError as e: # 捕获 ImportError 及其细节
    print(f"警告: 找不到核心功能模块 'image_scorer_supervised.py' 或其依赖项。个性化推荐评分功能将被跳过。错误: {e}")
    SCORER_MODULE_LOADED = False
# [2025-10-27] End image_scorer_supervised 导入


# [2025-06-10] 永远不要使用等待加载"networkidle"。
# [2025-06-10] 原有的功能不要乱改，应该多增加代码，少去删除以前的代码，没说让你改的地方别乱改，优先处理我说的问题。
# [2025-06-08] 代码一定是方便调试的，报错给我生成txt的log。存储信息用xlsx。log的txt和xlsx完成代码的时候要自动运行打开方便检查结果。
# [2025-06-08] 我提供的代码不要删除我的注释


def create_excel_report(image_data: List[Dict[str, Any]], base_filename="图片信息报告"):
    """
    Creates an Excel report from the collected image data with a timestamped filename.
    
    【核心修改点】：不再自动打开文件，以避免评分写入时的文件锁冲突。
    
    :param image_data: 包含图片信息的列表。
    :param base_filename: 报告的基础文件名。
    :return: 生成的 Excel 文件的绝对路径。
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_filename = f"{base_filename}_{timestamp}.xlsx"
    output_filepath = os.path.abspath(output_filename) # 获取绝对路径，方便后续调用

    df = pd.DataFrame(image_data)

    if df.empty:
        print("没有找到任何图片文件，将创建一个空的Excel文件。")
        # [2025-10-27] 更新: 增加所有可能的列名
        df = pd.DataFrame(columns=[
            "所在文件夹",
            "图片的绝对路径",
            "图片超链接",
            "stable diffusion的 ai图片的生成信息",
            "去掉换行符的生成信息",
            "正面提示词",
            "负面提示词",
            "其他设置",
            "正面提示词字数",
            "模型",
            "创建日期目录",
            "提取正向词的核心词",
            "TF-IDF区分度关键词(Top 10)" # 新增: TF-IDF 列
        ])
        # 注意: 'TF-IDF文件名后缀' 是临时列，不写入最终报告

    writer = pd.ExcelWriter(output_filepath, engine='openpyxl')
    # 过滤掉临时列 'TF-IDF文件名后缀'
    cols_to_write = [col for col in df.columns if col != TFIDF_SUFFIX_COLUMN]
    df[cols_to_write].to_excel(writer, index=False, sheet_name='图片信息')

    workbook = writer.book
    sheet = writer.sheets['图片信息']

    for col_idx, column_name in enumerate(cols_to_write): # 确保使用写入的列名列表
        # 强制设置所有列的宽度为20
        adjusted_width = 15
        
        sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

        if column_name == "图片超链接":
            for row_idx, cell_value in enumerate(df[column_name]):
                cell = sheet.cell(row=row_idx + 2, column=col_idx + 1)
                original_path = df["图片的绝对路径"].iloc[row_idx]
                cell.hyperlink = f"file:///{original_path}"
                cell.value = "点击查看原图"
                cell.font = Font(color=Color("0000FF"), underline="single")

    writer.close()
    print(f"数据已成功保存到 {output_filepath}")

    # @@    91-105,92-95   @@ 移除自动打开逻辑，避免文件锁冲突
    # 自动打开的逻辑移到主程序最后
    return output_filepath


# [2025-10-27] 新增函数: 用于调用 image_scorer_supervised 实现 Excel 添加两列 (难度系数: 1)
def add_scoring_columns_to_excel(excel_filepath: str) -> pd.DataFrame | None:
    """
    【难度系数: 1】调用 image_scorer_supervised 模块，对指定的 Excel 文件进行
    个性化推荐评分，并在原文件上添加两列评分结果。
    
    :param excel_filepath: 报告 Excel 文件的绝对路径。
    :return: 包含评分结果的 DataFrame, 失败返回 None。
    """
    if not SCORER_MODULE_LOADED:
        print("\n【评分功能跳过】'image_scorer_supervised.py' 模块未成功加载。")
        return None
        
    if not os.path.exists(excel_filepath):
        print(f"\n【评分功能失败】指定的 Excel 文件不存在: {excel_filepath}")
        return None

    print("\n--- 开始执行个性化推荐评分 (将添加'偏好定标分'和'个性化推荐预估评分'两列) ---")
    
    try:
        # 1. 初始化配置和评分器
        config = ScorerConfig()
        scorer = ImageScorer(config)
        
        # 2. 执行文件评分工作流（直接修改原文件）
        scorer.run_scoring_from_file(excel_filepath)
        
        print("--- 个性化推荐评分计算与保存完成 ---")
        
        # [2025-10-31] 新增: 重新读取已修改的Excel文件，并返回DataFrame
        # 这样主流程可以获取到更新后的评分列，用于文件名标记。
        print(f"重新读取已更新的Excel文件: {excel_filepath}...")
        updated_df = pd.read_excel(excel_filepath)
        return updated_df
        
    except Exception as e:
        print(f"【评分功能异常】调用 ImageScorer 失败，发生错误: {e}")
        return None
# [2025-10-27] End 新增函数


# [2025-10-31] 新增函数: 用于重命名图片文件，在文件名后添加评分标记
# @@    171-171,173-173   @@ 修改函数签名，返回新旧路径映射
def rename_images_with_score_tag(scored_df: pd.DataFrame, score_column_name: str, path_column_name: str) -> Dict[str, str]:
    """
    根据DataFrame中的'个性化推荐预估评分'列，对原始图片文件进行重命名，
    在文件名后添加评分标记 '@@@评分{分数}'，并移除/替换旧标记。
    
    【核心修改点】：返回旧路径到新路径的映射，用于更新内存中的 image_info。
    
    Args:
        scored_df (pd.DataFrame): 包含 '个性化推荐预估评分' 和 '图片的绝对路径' 的DataFrame。
        score_column_name (str): 评分列的名称。
        path_column_name (str): 图片绝对路径列的名称。
        
    Returns:
        Dict[str, str]: 旧路径到新路径的映射字典 {old_path: new_path}。
    """
    # 评分前缀需要在两边保持一致，这里硬编码以匹配 image_scorer_supervised.py 中的配置
    SCORE_PREFIX = "@@@评分"
    
    # [核心新增] 用于存储旧路径到新路径的映射
    path_map: Dict[str, str] = {}
    
    if scored_df is None or score_column_name not in scored_df.columns or path_column_name not in scored_df.columns:
        print("【评分标记重命名跳过】缺少必要的DataFrame或评分列。")
        return path_map # 返回空字典

    # 正则表达式用于匹配和移除旧的评分标记 (例如: @@@评分88)
    score_pattern = re.compile(rf'{re.escape(SCORE_PREFIX)}\d+$', re.IGNORECASE) 
    
    print("\n--- 开始执行图片文件评分标记重命名 (第二阶段 I/O: 评分Tag) ---")
    
    count_success = 0
    count_fail = 0
    
    # 迭代每一行数据进行重命名
    for index, row in tqdm(scored_df.iterrows(), total=len(scored_df), desc="重命名图片文件"):
        try:
            original_path = str(row[path_column_name])
            
            # 检查文件是否存在，防止重命名失败
            if not os.path.exists(original_path):
                 # 路径可能在标记阶段已被更新，但如果此时文件不存在，则跳过
                 continue
                 
            score = int(row[score_column_name])
            
            # 1. 解析路径
            directory = os.path.dirname(original_path)
            filename_ext = os.path.basename(original_path)
            base_name, ext = os.path.splitext(filename_ext)
            
            # 2. 移除旧的评分标记 (如果有)
            # 使用 sub('', base_name) 替换掉末尾的评分标记
            cleaned_base_name = score_pattern.sub('', base_name).strip().rstrip('_')
            
            # 3. 构建新的评分标记和文件名
            score_tag = f"{SCORE_PREFIX}{score}"
            
            # 新的文件名 = 清理旧评分后的文件名 + 新评分标记
            new_base_name = f"{cleaned_base_name}{score_tag}" 
            
            new_path = os.path.join(directory, new_base_name + ext)
            
            # 4. 执行重命名
            if original_path.lower() != new_path.lower():
                os.rename(original_path, new_path)
                count_success += 1
                
                # [核心修改] 记录路径映射
                path_map[original_path] = new_path
            else:
                # 文件名未发生变化 (评分相同或无评分标记)
                pass 

        except Exception as e:
            # log_error(f"【重命名失败】文件: {original_path}，错误: {e}") # 避免在 tqdm 中打印大量日志
            count_fail += 1

    print(f"--- 图片评分标记重命名完成 ---")
    print(f"总图片数: {len(scored_df)}, 成功重命名: {count_success}, 失败: {count_fail}")
    
    return path_map # 返回路径映射

# @@    267-272,272-277   @@ 统一分类选项变量
if __name__ == "__main__":
    print("""
    ======================================================================
    ||                 AI 图片信息提取与文件两级分类工具                ||
    ======================================================================
    
    核心功能流程:
    1. 【输入】获取用户指定的图片根目录（或使用默认路径）。
    2. 【归档】如果选择的目录是历史归档目录或项目根目录，自动将项目中的日期文件夹归档到历史目录。
    3. 【扫描】**[更新]** 特殊文件夹扫描整个项目以获取评分基准，非特殊文件夹只扫描输入路径。
    4. 【选择】**[新增]** 用户选择：**1. 标记/评分重命名**、**2. 分类**、**3. 只生成报告**。
    5. 【报告/分类】根据选择执行对应操作。
    
    ----------------------------------------------------------------------
    """) # 打印核心功能介绍

    # [2025-10-31] 项目根目录路径（即 SD 的 outputs 目录）
    PROJECT_FOLDER_PATH = r"C:\stable-diffusion-webui\outputs\txt2img-images"
    # [2025-10-30] 归档目标文件夹路径 / 默认扫描路径
    DEFAULT_FOLDER_PATH = r"C:\stable-diffusion-webui\outputs\txt2img-images\历史" 
    
    # [2025-10-22] 归档触发路径的标准化绝对路径 (即 '历史' 文件夹)
    ARCHIVE_TRIGGER_PATH = os.path.abspath(DEFAULT_FOLDER_PATH) 
    # [2025-10-31] 项目根目录的绝对路径 (用于扫描语料和安全检查)
    RISKY_ROOT_DIR_ABS = os.path.abspath(PROJECT_FOLDER_PATH)
    
    # [2025-10-31] 模型基准评分文件夹（用于 TF-IDF 和模型训练的语料库来源）
    # C:\stable-diffusion-webui\outputs\txt2img-images\精选       80分
    # C:\stable-diffusion-webui\outputs\txt2img-images\超级精选   85分
    # C:\stable-diffusion-webui\outputs\txt2img-images\特殊画风   90分
    # C:\stable-diffusion-webui\outputs\txt2img-images\超绝       95分

    # @@    318-318,335-373   @@ 将功能选择菜单提前到路径输入之前
    # --- 核心修改: 功能选择菜单提前 (最低改动点) ---
    print("\n========================================================")
    print("||              🎯 扫描结果处理操作选择              ||")
    print("========================================================")
    print("1. **执行标记 & 评分重命名** & 生成报告")
    print("2. **执行文件分类** & 生成报告 (将文件移动到新位置，**高风险操作**)")
    print("3. **只生成报告** (不打标, 不重命名, 不分类)")
    print("4. 退出程序")
    
    process_choice = input("请选择要执行的操作 (1, 2, 3, 4): ").strip()
    
    if process_choice == "4":
        print("程序已退出。")
        exit()
        
    if process_choice not in ["1", "2", "3"]:
        print(f"\n选择 '{process_choice}' 无效，程序结束。")
        exit()
    # --- 核心修改结束 ---

    # 路径输入现在只在选择了有效操作后才执行
    folder_to_scan = input(f"请输入要扫描的主文件夹路径 (回车使用默认路径: {DEFAULT_FOLDER_PATH}): ").strip()

    if not folder_to_scan:
        folder_to_scan = DEFAULT_FOLDER_PATH # 如果输入为空，则使用默认路径
        print(f"已使用默认文件夹路径: {folder_to_scan}")

    if not os.path.isdir(folder_to_scan):
        print(f"错误: 文件夹 '{folder_to_scan}' 不存在。请提供一个有效的文件夹路径。")
        log_error(f"用户输入的文件夹 '{folder_to_scan}' 不存在。") # 记录文件夹不存在的错误
        
        
    else:
        # 获取用户输入文件夹的绝对路径，用于判断是否触发归档和安全检查
        current_scan_path_abs = os.path.abspath(folder_to_scan)
        
        # [2025-10-31] 判断是否为项目文件夹或归档文件夹
        IS_SPECIAL_FOLDER = (current_scan_path_abs == RISKY_ROOT_DIR_ABS) or (current_scan_path_abs == ARCHIVE_TRIGGER_PATH)
        
        # --- 归档逻辑：仅在特殊路径时触发归档 ---
        if IS_SPECIAL_FOLDER:
            print(f"\n--- 归档操作 ---")
            print(f"当前路径 '{folder_to_scan}' 匹配特殊文件夹，执行日期文件夹归档。")
            # 归档源基目录 (base_source_dir) 是项目根目录 (PROJECT_FOLDER_PATH)
            archive_target_dir = ARCHIVE_TRIGGER_PATH # 目标一定是 历史 文件夹
            base_source_dir = RISKY_ROOT_DIR_ABS # 源一定是项目根目录

            # 执行归档
            archive_date_folders(base_source_dir, archive_target_dir) 
        else:
            print("\n--- 归档操作 ---")
            print(f"当前路径 '{folder_to_scan}' 为非特殊文件夹，跳过日期文件夹归档。")
        # --- 归档逻辑结束 ---

        # 1. 扫描文件夹并生成报告
        print(f"\n--- 扫描操作 ---")
        if IS_SPECIAL_FOLDER:
            # 特殊文件夹：扫描整个项目根目录，以获取所有基准评分文件夹的语料
            folder_to_scan_actual = RISKY_ROOT_DIR_ABS
            print(f"检测到特殊文件夹，扫描范围扩大到整个项目文件夹: {folder_to_scan_actual}")
        else:
            # 非特殊文件夹：仅扫描用户输入的文件夹
            folder_to_scan_actual = folder_to_scan
            print(f"非特殊文件夹，仅扫描用户指定文件夹: {folder_to_scan_actual}")
            
        # 调用从 image_scanner 导入的函数
        image_info = get_image_info(folder_to_scan_actual) 
        
        if not image_info:
            print("没有扫描到任何图片信息，程序结束。")
            exit()
            
        # 初始化报告路径和 DataFrame
        report_path = None
        df_for_process = pd.DataFrame(image_info)
        
        # --- 2. 逻辑分支 1: 标记 & 评分重命名 (原有的 TF-IDF, 标记, 评分逻辑) ---
        if process_choice == "1":
            print("\n>>> 您选择了：1. 执行标记 & 评分重命名 & 生成报告 <<<")
            
            # --- 2. TF-IDF 区分度分析功能 (仅在特殊文件夹下运行 TF-IDF 分析) ---
            if IS_SPECIAL_FOLDER and TFIDF_TARGET_COLUMN in df_for_process.columns:
                print("\n--- 检测到特殊文件夹，开始执行 TF-IDF 区分度分析 (多进程加速) ---")
                
                # 1) 数据预处理 
                original_tags_series = df_for_process[TFIDF_TARGET_COLUMN]
                corpus, cleaned_tags_series = preprocess_tags(original_tags_series)
                
                if corpus:
                    # 2) 计算并提取 Top N 关键词
                    distinctive_features_excel, distinctive_tag_lists = calculate_and_extract_tfidf(
                        df_for_process, 
                        corpus, 
                        cleaned_tags_series, 
                        TFIDF_TOP_N_FEATURES
                    )
                    
                    # 3) 将结果添加到 DataFrame
                    if len(distinctive_features_excel) == len(df_for_process):
                        df_for_process[TFIDF_NEW_COLUMN_NAME] = distinctive_features_excel # 写入 Excel 列
                        
                        # [核心新增] 将 TF-IDF 关键词列表转换为文件名后缀字符串并存入临时列
                        df_for_process[TFIDF_SUFFIX_COLUMN] = [
                            format_tfidf_tags_for_filename(tag_list) for tag_list in distinctive_tag_lists
                        ]
                    else:
                        print("警告: TF-IDF 结果长度不匹配，跳过添加新列。")
                        df_for_process[TFIDF_NEW_COLUMN_NAME] = "结果长度不匹配"
                        
                else:
                    print("警告: TF-IDF 语料库为空，跳过计算。")
                    df_for_process[TFIDF_NEW_COLUMN_NAME] = "语料为空"
                
                print("--- TF-IDF 区分度分析完成 ---")
            else:
                if not IS_SPECIAL_FOLDER:
                    print("\n--- 非特殊文件夹，跳过 TF-IDF 分析 ---")
            # --- End TF-IDF 功能 ---

            # 5. 将处理后的 DataFrame 转换回 list of dicts，包含新增的 TF-IDF 列
            # 这一步是为了 tag_files_by_prompt 函数的兼容性
            image_info_for_tagging = df_for_process.to_dict('records')

            # 6. 执行文件名标记操作 (仅在特殊文件夹下运行自定义和TF-IDF文件名标记)
            # 标记后缀顺序: 文件名 -> 自定义tag -> TF-IDF自动标记tag
            if IS_SPECIAL_FOLDER:
                print(f"\n当前文件名标记关键词列表 (从 filename_tagger 导入): {TAGGING_KEYWORDS} (定界符: ___ )")
                print("\n--- 检测到特殊文件夹，执行文件名标记 (第一阶段 I/O: 自定义Tag和TF-IDF Tag) ---")
                # 将 TF-IDF 临时列名传入，实现二次附加后缀
                image_info_for_tagging = tag_files_by_prompt(image_info_for_tagging, TAGGING_KEYWORDS, tfidf_suffix_col=TFIDF_SUFFIX_COLUMN)
                # 重新转换为 DataFrame 以便生成报告和评分
                df_for_process = pd.DataFrame(image_info_for_tagging)
            else:
                print("\n--- 非特殊文件夹，跳过文件名标记 (自定义Tag/TF-IDF Tag 默认不启用) ---")
            
            # 7. 生成报告 (此时 df_for_process 已经是最新的，且路径已更新)
            print("\n--- 报告生成 ---")
            report_path = create_excel_report(df_for_process.to_dict('records')) # 接收返回的路径
            
            # [2025-10-31] 新增变量: 用于存储评分后的 DataFrame 和路径映射
            scored_df: Union[pd.DataFrame, None] = None 
            path_update_map: Dict[str, str] = {} # 用于存储旧路径 -> 新路径的映射
            
            # --- 评分功能调用逻辑 ---
            if report_path: # 确保报告文件生成成功
                print("\n--- 个性化推荐评分选项 (影响文件重命名) ---")
                
                # [2025-10-31] 调整提示：非特殊文件夹回车默认跳过评分
                prompt_options = "1/回车: 添加评分, 2/no: 跳过评分" if IS_SPECIAL_FOLDER else "1: 添加评分, 2/回车/no: 跳过评分"
                score_choice = input(f"是否为生成的 Excel 报告添加个性化推荐评分列？({prompt_options}): ").strip().lower()
                
                # [2025-10-31] 评分触发条件: 手动输入 "1" 或 特殊文件夹且输入为空
                enable_scoring = (score_choice == "1") or (score_choice == "" and IS_SPECIAL_FOLDER)
                
                if enable_scoring:
                    print("\n您选择了添加个性化推荐评分。")
                    # 调用评分函数，它现在会修改 Excel 文件并返回评分后的 DataFrame
                    scored_df = add_scoring_columns_to_excel(report_path) 
                    
                    # [2025-10-31] 新增: 评分标记重命名也仅在特殊文件夹下执行
                    if IS_SPECIAL_FOLDER:
                        if scored_df is not None and '个性化推荐预估评分' in scored_df.columns:
                            PREDICTED_SCORE_COLUMN = '个性化推荐预估评分'
                            # 确保找到正确的路径列名
                            PATH_COLUMN_NAME = '图片的绝对路径' 
                            
                            # @@    538-538,541-541   @@ 执行评分标记重命名 (返回路径映射)
                            path_update_map = rename_images_with_score_tag(scored_df, PREDICTED_SCORE_COLUMN, PATH_COLUMN_NAME)
                        
                elif score_choice in ["2", "no", ""]: # 统一跳过逻辑
                    print("\n您选择了跳过个性化推荐评分。")
                else:
                    print("\n评分输入无效，默认跳过个性化推荐评分。")
            # --- 评分功能调用逻辑结束 ---
            
            # [核心新增] 使用路径映射更新 image_info 列表 (解决分类失败问题)
            if path_update_map:
                print("\n--- 更新内存中的文件路径 (解决分类找不到文件问题) ---")
                # 更新 image_info_for_tagging 列表
                for item in image_info_for_tagging:
                    old_path = item["图片的绝对路径"]
                    if old_path in path_update_map:
                        new_path = path_update_map[old_path]
                        item["图片的绝对路径"] = new_path
                        
            # 将最终处理过的 image_info_for_tagging 赋值给 image_info，以便后续流程（如果有）
            image_info = image_info_for_tagging


        # --- 3. 逻辑分支 2: 文件分类 (原有的分类逻辑) ---
        elif process_choice == "2":
            print("\n>>> 您选择了：2. 执行文件分类 & 生成报告 <<<")
            
            # 7. 生成报告 (在分类前生成，报告中路径为分类前路径)
            print("\n--- 报告生成 (分类前) ---")
            report_path = create_excel_report(image_info) 
            
            # 3. 定义默认分类关键词 (用于分类文件夹，不是用于文件名标记)
            default_keywords = "skeleton,penis,pussy,nipple,vagina,censor,nude,green_hair,blue_hair,red_hair,purple_hair,yellow_hair,pink_hair,white_hair,grey_hair,brown_hair,black_hair,blonde_hair,aqua_hair"
            keyword_list = [kw.strip() for kw in default_keywords.split(',')]
            print(f"\n当前默认分类关键词列表 (用于文件夹分类): {default_keywords}")

            # 8. 分类选项
            print("\n--- 分类选项 ---")
            choice = None # 初始化 choice
            
            if current_scan_path_abs == RISKY_ROOT_DIR_ABS:
                # 匹配到项目根目录，强制关闭分类 (选项 no)
                choice = "no" # 使用 'no' 统一逻辑
                print(f"【!! 严重警告 !!】路径 '{folder_to_scan}' 是项目根目录 ('{RISKY_ROOT_DIR_ABS}')。")
                print("已输入项目文件夹，不进行自动分类，只生成表格。")
                log_error(f"阻止了在危险根目录 '{folder_to_scan}' 上执行分类操作。")
                
            elif current_scan_path_abs == ARCHIVE_TRIGGER_PATH:
                # 匹配到默认的“历史”文件夹，询问用户是否自动分类
                print(f"路径 '{folder_to_scan}' 匹配默认归档文件夹。")
                
                # 询问用户，允许回车或 'yes' 启用，或 '2/no' 跳过
                classification_prompt = "是否自动执行分类操作？(yes/回车: 自动分类, 2/no: 跳过): "
                user_input = input(classification_prompt).strip().lower()

                if user_input in ["yes", "1", ""]:
                    choice = "yes"
                    print("已选择自动分类。")
                elif user_input in ["2", "no"]:
                    choice = "no"
                    print("已选择跳过分类。")
                else:
                    choice = "no"
                    print("输入无效，默认跳过分类。")
                
            else:
                # 非特殊路径，提示用户输入 (强制输入 'yes' 或 '2/no')
                print("【!!! 高风险区域 - 非特殊文件夹 !!!】")
                print("请认真考虑是否要执行自动分类操作！！！")
                # 强制要求输入 'yes' 或 '2/no'
                classification_prompt = "是否自动执行分类操作？(yes: 自动分类, 2/no: 仅生成报告不分类): "
                choice = input(classification_prompt).strip().lower()
            
            # --- 执行分类判断 ---
            if choice in ["1", "yes"]:
                print("\n您选择了自动分类。")
                # 调用从 file_categorizer 导入的函数
                categorize_images(image_info, keyword_list, folder_to_scan)
            elif choice in ["2", "no"]: # 接受 "2" 或 "no"
                print("\n您选择了不自动分类，已生成报告。")
            else:
                # 捕获无效输入
                print("\n输入无效，默认不执行分类，已生成报告。")


        # --- 4. 逻辑分支 3: 只生成报告 ---
        elif process_choice == "3":
            print("\n>>> 您选择了：3. 只生成报告 (不打标, 不重命名, 不分类) <<<")
            
            # 直接使用原始扫描结果生成报告
            print("\n--- 报告生成 ---")
            report_path = create_excel_report(image_info) 


        # --- 5. 最终报告打开逻辑 (无论选择哪个流程，报告生成后都尝试打开) ---
        if report_path:
            # 自动打开报告文件
            try:
                if os.name == 'nt': 
                    subprocess.run(['start', report_path], shell=True, check=True)
                elif os.uname().sysname == 'Darwin': 
                    subprocess.run(['open', report_path], check=True)
                else: 
                    subprocess.run(['xdg-open', report_path], check=True)
                print(f"\n--- 所有处理完成，尝试自动打开报告文件: {report_path} ---")
            except Exception as e:
                print(f"\n--- 报告文件打开失败 ---")
                print(f"错误: 无法自动打开 '{report_path}'。请手动打开。错误信息: {e}")
        
        # @@    829-831,833-833   @@ 移除无效选择的打印，因为在功能选择时已处理